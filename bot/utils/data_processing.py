import os
import datetime
import logging
import warnings
from typing import List, Dict, Any, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from bot.utils.functions import (
    fill_cadastr_num, 
    fill_area,
    get_coords_from_cadastral_number, 
    load_constants,
    convert_time,
    fill_rent_period,
    get_additional_data,
    get_coords_batch,
    get_additional_data_batch
)


# Игнорируем предупреждения
warnings.filterwarnings('ignore')

logger = logging.getLogger(__name__)


def prepare_data_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Подготавливает данные для Excel файла"""
    # Выбираем и переименовываем нужные колонки
    columns_mapping = {
        'id': 'ID лота',
        'lotName': 'Наименование',
        'lotDescription': 'Описание',
        'category': 'Категория',
        'biddType': 'Тип торгов',
        'biddForm': 'Форма торгов',
        'subject': 'Субъект РФ',
        'address': 'Адрес',
        'cadastral_number': 'Кадастровый номер',
        'area': 'Площадь (кв.м)',
        'priceMin': 'Начальная цена (руб)',
        'priceFin': 'Конечная цена (руб)',
        'deposit': 'Задаток (руб)',
        'priceStep': 'Шаг аукциона (руб)',
        'rent_period': 'Срок аренды',
        'biddEndTime': 'Дата окончания приема заявок',
        'auction_start_date': 'Дата проведения аукциона',
        'bidd_start_date': 'Дата начала приема заявок',
        'lotStatus': 'Статус',
        'coordinates_xy': 'Координаты лота',
        'link': 'Ссылка на лот',
        'auction_link': 'Ссылка на аукцион',
        'lotImages': 'Изображения',
        'files': 'Документы'
    }
    
    # Создаем новый DataFrame с нужными колонками
    result_df = pd.DataFrame()
    
    for new_col, old_col in columns_mapping.items():
        if new_col in df.columns:
            result_df[old_col] = df[new_col]
    
    # Форматируем даты
    date_columns = ['Дата окончания приема заявок', 'Дата проведения аукциона', 'Дата начала приема заявок']
    for col in date_columns:
        if col in result_df.columns:
            result_df[col] = pd.to_datetime(result_df[col]).dt.strftime('%d.%m.%Y %H:%M')
    
    # Форматируем числовые значения
    numeric_columns = ['Начальная цена (руб)', 'Задаток (руб)', 'Шаг аукциона (руб)']
    for col in numeric_columns:
        if col in result_df.columns:
            result_df[col] = result_df[col].apply(lambda x: f"{x:,.2f}".replace(',', ' ') if pd.notnull(x) else '')
    
    # Преобразуем конечную цену в целое число
    if 'Конечная цена (руб)' in result_df.columns:
        result_df['Конечная цена (руб)'] = result_df['Конечная цена (руб)'].apply(
            lambda x: int(x) if pd.notnull(x) else ''
        )
    
    # Форматируем координаты
    if 'Координаты лота' in result_df.columns:
        result_df['Координаты лота'] = result_df['Координаты лота'].apply(
            lambda x: f"{x[0]}, {x[1]}" if isinstance(x, list) and len(x) == 2 else ''
        )
    
    # Форматируем ссылки на изображения
    if 'Изображения' in result_df.columns:
        result_df['Изображения'] = result_df['Изображения'].apply(
            lambda x: '\n'.join([f"https://torgi.gov.ru/new/file-store/v1/{img}?disposition=inline" for img in x]) if isinstance(x, list) and all(isinstance(img, str) for img in x) else ''
        )
    
    # Форматируем ссылки на документы
    if 'files' in df.columns:
        result_df['Документы'] = df['files'].apply(
            lambda x: '\n'.join([f"{name}: {url}" for name, url in x]) if isinstance(x, list) and all(isinstance(item, tuple) and len(item) == 2 for item in x) else ''
        )
    
    return result_df


def format_excel(wb: Workbook, sheet_name: str) -> None:
    """Форматирует Excel файл"""
    ws = wb[sheet_name]
    
    # Устанавливаем ширину колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = min(adjusted_width, 50)
    
    # Форматируем заголовки
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Форматируем данные
    data_alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_alignment
    
    # Добавляем границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    # Закрепляем заголовки
    ws.freeze_panes = "A2"


def data_processing(data: List[Dict[Any, Any]], selected_subjects: List[str], selected_statuses: List[str], config=None) -> str:
    """Обрабатывает данные и создает Excel файл"""
    logger.info("Начинаю обработку данных...")
    
    # Загружаем константы
    subjects_data, _, statuses_data = load_constants()
    
    # Создаем DataFrame из полученных данных
    df = pd.DataFrame(data)
    
    if df.empty:
        logger.error("Нет данных для обработки")
        return None
    
    # Обрабатываем данные
    logger.info("Обрабатываю данные...")
    
    # Добавляем информацию о субъекте
    if 'subjectRFCode' in df.columns:
        df['subject'] = df['subjectRFCode'].apply(
            lambda x: next((sub['name'] for sub in subjects_data if sub['subjectRFCode'] == str(x)), "Неизвестный субъект")
        )
    else:
        logger.error("В данных отсутствует поле subjectRFCode")
    
    if 'lotStatus' in df.columns:
        df['lotStatus'] = df['lotStatus'].apply(
            lambda x: next((status['name'] for status in statuses_data if status['code'] == str(x)), "Неизвестный статус")
        )
    else:
        logger.error("В данных отсутствует поле lotStatus")

    if 'attributes' in df.columns:
        df['rent_period'] = df['attributes'].apply(fill_rent_period)

    if 'characteristics' in df.columns:
        df['area'] = df['characteristics'].apply(fill_area)

    df['cadastral_number'] = df.apply(lambda x: fill_cadastr_num(x['characteristics'], x['lotDescription']), axis=1)
        
    # Обрабатываем изображения
    if 'lotImages' in df.columns:
        try:
            df['lotImages'] = df['lotImages'].apply(lambda x: '\n'.join([f'https://torgi.gov.ru/new/file-store/v1/{img}?disposition=inline' for img in x]))
        except Exception as e:
            logger.error(f"Ошибка при обработке изображений: {e}")
            df['lotImages'] = [[]]  # Устанавливаем пустой список, чтобы избежать ошибок
    
    df['link'] = df['id'].apply(lambda x: f'https://torgi.gov.ru/new/public/lots/lot/{x}') 

    # Рассчитываем координаты, если это требуется
    if config and config.processing.calculate_coordinates and 'cadastral_number' in df.columns:
        logger.info("Рассчитываю координаты по кадастровым номерам...")
        
        # Получаем уникальные непустые кадастровые номера
        unique_cadastral_numbers = df['cadastral_number'].dropna().unique().tolist()
        
        if unique_cadastral_numbers:
            # Ограничиваем количество запросов в зависимости от размера данных
            workers = min(5, max(2, len(unique_cadastral_numbers) // 20))
            logger.info(f"Будет использовано {workers} параллельных потоков для запросов")
            
            # Получаем координаты параллельно с контролем скорости запросов
            coords_dict = get_coords_batch(
                unique_cadastral_numbers, 
                max_workers=workers,
                retry_interval=3,
                rate_limit_delay=1.0  # Увеличиваем задержку между запросами
            )
            
            # Применяем результаты к DataFrame через map
            coordinates_map = {cad_num: coords for cad_num, coords in coords_dict.items()}
            df['coordinates'] = df['cadastral_number'].map(
                lambda x: coordinates_map.get(x, np.nan) if pd.notnull(x) and x else np.nan
            )
        else:
            df['coordinates'] = np.nan
            
        # Разделяем координаты и адрес
        df['coordinates_xy'] = df['coordinates'].apply(
            lambda x: x[0][::-1] if isinstance(x, tuple) and len(x) > 0 and x[0] is not np.nan else np.nan
        )
        df['address'] = df['coordinates'].apply(
            lambda x: x[1] if isinstance(x, tuple) and len(x) > 1 else np.nan
        )
        df['yandex_map_link'] = df['coordinates_xy'].apply(
            lambda x: f"https://yandex.ru/maps/?text={x[0]},{x[1]}" if isinstance(x, list | tuple) and len(x) > 0 and x is not np.nan else np.nan
        )

    # Преобразуем типы данных
    if 'biddType' in df.columns:
        df['biddType'] = df['biddType'].apply(lambda x: x['name'] if isinstance(x, dict) and 'name' in x else x)
    
    if 'biddForm' in df.columns:
        df['biddForm'] = df['biddForm'].apply(lambda x: x['name'] if isinstance(x, dict) and 'name' in x else x)
    
    if 'category' in df.columns:
        df['category'] = df['category'].apply(lambda x: x['name'] if isinstance(x, dict) and 'name' in x else x)
    

    # Преобразуем даты
    try:
        if 'biddEndTime' in df.columns:
            df['biddEndTime'] = df.apply(lambda x: convert_time(x['biddEndTime'], x.get('timezoneOffset', 0)), axis=1)
        if 'createDate' in df.columns:
            df['createDate'] = df.apply(lambda x: convert_time(x['createDate'], x.get('timezoneOffset', 0)), axis=1)
        if 'auction_start_date' in df.columns:
            df['auction_start_date'] = df.apply(lambda x: convert_time(x['auction_start_date'], x.get('timezoneOffset', 0)), axis=1)
        if 'bidd_start_date' in df.columns:
            df['bidd_start_date'] = df.apply(lambda x: convert_time(x['bidd_start_date'], x.get('timezoneOffset', 0)), axis=1)
    except Exception as e:
        logger.error(f'Ошибка в преобразовании времени: {e}')
    
    try:
        logger.info('Начинаю собирать дополнительные данные об объекте...')
        
        # Получаем уникальные непустые ID лотов
        unique_lot_ids = df['id'].dropna().unique().tolist()
        
        if unique_lot_ids:
            # Получаем дополнительные данные параллельно
            additional_data_dict = get_additional_data_batch(unique_lot_ids, max_workers=10)
            
            # Создаем временные колонки для данных
            data_columns = ['auction_start_date', 'bidd_start_date', 'auction_link', 
                             'price_step', 'deposit_price', 'files', 'permitted_use']
            
            # Преобразуем словарь результатов в DataFrame для удобного соединения
            additional_df = pd.DataFrame([
                [lot_id] + list(data)
                for lot_id, data in additional_data_dict.items()
            ], columns=['id'] + data_columns)
            
            # Объединяем с основным DataFrame
            df = df.drop(columns=[col for col in data_columns if col in df.columns]).merge(
                additional_df, on='id', how='left'
            )
        else:
            for col in ['auction_start_date', 'bidd_start_date', 'auction_link', 
                       'price_step', 'deposit_price', 'files', 'permitted_use']:
                df[col] = np.nan
        
        logger.info('Дополнительные данные собраны!')
        
        if 'files' in df.columns:
            df['files'] = df['files'].apply(
                lambda x: '\n'.join([f"{name}: {url}" for name, url in x]) if isinstance(x, list) and all(isinstance(item, tuple) and len(item) == 2 for item in x) else ''
            )
    except Exception as e:
        logger.error(f'Ошибка в получении дополнительных данных: {e}')

    # Удаляем ненужные колонки
    # columns_to_drop = ['characteristics', 'attributes', 'subjectRFCode']
    # df = df.drop(columns=[col for col in columns_to_drop if col in df.columns]).reset_index(drop=True)
    
    df.rename(columns={
        'priceMin': 'price_min',
        'priceFin': 'price_fin',
        'biddType': 'bidd_type',
        'biddForm': 'bidd_form',
        'lotStatus': 'lot_status',
        'biddEndTime': 'bidd_end_date',
        'lotImages': 'lot_images',
        'lotName': 'lot_name',
        'lotDescription': 'lot_description',
    }, inplace=True)

    df_base_columns = [
        'id', 'link', 'lot_name', 'lot_description', 'category', 'subject', 'permitted_use', 'lot_status', 
        'bidd_type', 'bidd_form', 'bidd_start_date', 'bidd_end_date', 'auction_start_date', 'auction_link',
        'deposit_price','price_min', 'price_step', 'price_fin', 'rent_period', 'area', 'cadastral_number', 'lot_images', 'files'
    ]
    df_coords_columns = ['coordinates_xy', 'address', 'yandex_map_link']


    df = df[df_base_columns + df_coords_columns].reset_index(drop=True) if 'coordinates_xy' in df.columns else df[df_base_columns].reset_index(drop=True)

    # Подготавливаем данные для Excel
    # excel_df = prepare_data_for_excel(df)
    
    # Создаем директорию для результатов, если она не существует
    results_path = os.path.join('data', 'results')
    os.makedirs(results_path, exist_ok=True)
    
    # Формируем имя файла
    time_now = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Получаем названия субъектов
    subject_names = []
    for subject_code in selected_subjects:
        subject_name = next((sub['name'] for sub in subjects_data if sub['code'] == subject_code), None)
        if subject_name:
            subject_names.append(subject_name.replace(" ", "_"))
    
    # Формируем имя файла
    subjects_str = "-".join(subject_names) if len(subject_names) <= 2 else f"{len(subject_names)}_субъектов"
    statuses_str = "-".join(selected_statuses) if len(selected_statuses) <= 2 else f"{len(selected_statuses)}_статусов"
    filename = f"TORGI_{subjects_str}_{statuses_str}_{time_now}.xlsx"
    
    # Полный путь к файлу
    file_path = os.path.join(results_path, filename)
    
    # Создаем Excel файл
    logger.info(f"Создаю Excel файл: {file_path}")
    
    # Сохраняем данные в Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Данные', index=False)
        workbook = writer.book
        format_excel(workbook, 'Данные')
    
    logger.info(f"Excel файл успешно создан: {file_path}")
    
    return file_path


def process_images(images_data):
    """Обрабатывает данные изображений"""
    try:
        if not isinstance(images_data, list):
            return []
        
        result = []
        for img in images_data:
            if isinstance(img, dict) and 'fileId' in img:
                result.append(img['fileId'])
            elif isinstance(img, str):
                result.append(img)
        return result
    except Exception as e:
        logger.error(f"Ошибка в process_images: {e}, тип данных: {type(images_data)}")
        return []
